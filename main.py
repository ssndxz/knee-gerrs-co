import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
import seaborn as sns
import os
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime, timedelta

DB_NAME = "dataviz"
DB_USER = "postgres"
DB_PASS = "1234"
DB_HOST = "localhost"
DB_PORT = "5432"
QUERIES_FILE = "queries.sql"
CHARTS_DIR = "charts"
EXPORTS_DIR = "exports"

engine = create_engine(f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}")


def load_queries(filename):
    try:
        with open(filename, "r", encoding="utf-8") as f:
            content = f.read()
    except FileNotFoundError:
        print(f"Error: {filename} not found")
        return {}
    queries = {}
    current_name = None
    buffer = []
    for line in content.splitlines():
        if line.strip().startswith("-- name:"):
            if current_name and buffer:
                queries[current_name] = "\n".join(buffer).strip()
                buffer = []
            current_name = line.replace("-- name:", "").strip()
        else:
            buffer.append(line)
    if current_name and buffer:
        queries[current_name] = "\n".join(buffer).strip()
    return queries


def run_query(query, name):
    try:
        df = pd.read_sql(text(query), engine)
        print(f"[Debug] {name} columns: {df.columns.tolist()}")
        if name != "events_by_date":
            csv_path = os.path.join(EXPORTS_DIR, f"{name}.csv")
            df.to_csv(csv_path, index=False, encoding="utf-8")
            print(f"[Query] {name}: {len(df)} rows (saved {csv_path})")
        else:
            print(f"[Query] {name}: {len(df)} rows (not saved to CSV)")
        return df
    except Exception as e:
        print(f"[Query Error] {name}: {e}")
        return pd.DataFrame()


def generate_chart(df, chart_type, filename, x_col, y_col, title, xlabel, ylabel):
    if df.empty:
        print(f"[Chart Error] {filename}: No data to plot")
        return
    plt.figure(figsize=(10, 6))
    sns.set_style("whitegrid")

    if chart_type == "pie":
        plt.pie(df[y_col], labels=df[x_col], autopct='%1.1f%%')
    elif chart_type == "bar":
        plt.bar(df[x_col], df[y_col])
    elif chart_type == "barh":
        plt.barh(df[x_col], df[y_col])
    elif chart_type == "line":
        plt.plot(df[x_col], df[y_col], marker='o')
    elif chart_type == "hist":
        plt.bar(df[x_col], df[y_col], edgecolor='black')
        print(f"[Debug] {filename} - Plotted values: {dict(zip(df[x_col], df[y_col]))}")
    elif chart_type == "scatter":
        plt.scatter(df[x_col], df[y_col])
        for i, name in enumerate(df[x_col]):
            plt.text(df[x_col][i], df[y_col][i], name, fontsize=8)

    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.tight_layout()

    path = os.path.join(CHARTS_DIR, filename)
    plt.savefig(path)
    plt.close()
    print(f"[Chart] {filename} → {len(df)} rows, type={chart_type}, shows: {title}")


def plotly_time_slider(df):
    if df.empty or "date" not in df.columns:
        print("[Plotly Error] No valid data or Date column for time slider")
        return
    df["date"] = pd.to_datetime(df["date"])
    
    min_date = pd.to_datetime("2023-01-01")
    max_date = pd.to_datetime("2025-12-31")
    date_range = pd.date_range(start=min_date, end=max_date, freq='D')

    df_expanded = pd.DataFrame({'date': date_range})
    df_expanded = df_expanded.merge(df.groupby('date').agg({'eventcount': 'sum', 'name': 'first', 'genrename': 'first'}).reset_index(), on='date', how='left')
    df_expanded['eventcount'].fillna(0, inplace=True)
    df_expanded['name'].fillna('No Events', inplace=True)
    df_expanded['genrename'].fillna('None', inplace=True)
    df_expanded['year'] = df_expanded['date'].dt.year
    df_expanded['day_of_year'] = df_expanded['date'].dt.dayofyear
    
    fig = px.scatter(df_expanded, 
                     x="day_of_year", 
                     y="eventcount", 
                     color="genrename",
                     size="eventcount",
                     animation_frame="year",
                     range_x=[1, 366],
                     title="Interactive: Daily Event Counts by Year",
                     labels={"day_of_year": "Day of Year", "eventcount": "Event Count"})
    
 
    fig.layout.updatemenus[0].buttons[0].args[1]["frame"]["duration"] = 1000  # 1s per year
    fig.layout.updatemenus[0].buttons[0].args[1]["frame"]["redraw"] = True
    
    fig.show()
    print("[Plotly] Displayed interactive scatter plot with year-by-year slider")


def export_formatted_excel(dataframes_dict, filename):
    filepath = os.path.join(EXPORTS_DIR, filename)
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet, df in dataframes_dict.items():
            if sheet != "events_by_date":
                df.to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(filepath)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            if all(isinstance(cell.value, (int, float, type(None))) for cell in col):
                col_letter = col[0].column_letter
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)
    wb.save(filepath)
    total_rows = sum(len(df) for df in dataframes_dict.values() if df is not None and df.empty is False)
    print(f"[Export] Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")


def main():
    os.makedirs(CHARTS_DIR, exist_ok=True)
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    queries = load_queries(QUERIES_FILE)
    print(f"[Info] Loaded queries: {list(queries.keys())}")

    dataframes = {}

    # Bar Chart
    df1 = run_query(queries["top_rated_events"], "top_rated_events")
    generate_chart(df1, "bar", "top_rated_events.png",
                   x_col="name", y_col="avgrate",
                   title="Top 5 Highest Rated Events",
                   xlabel="Event Name", ylabel="Average Rating")
    dataframes["top_rated_events"] = df1

    # Pie Chart
    df2 = run_query(queries["events_per_location"], "events_per_location")
    generate_chart(df2, "pie", "events_per_location.png",
                   x_col="name", y_col="eventcount",
                   title="Distribution of Events by Location",
                   xlabel="", ylabel="Number of Events")
    dataframes["events_per_location"] = df2

    # Horizontal Bar Chart
    df3 = run_query(queries["artists_most_events"], "artists_most_events")
    generate_chart(df3, "barh", "artists_most_events.png",
                   x_col="name", y_col="eventcount",
                   title="Artists with Most Events",
                   xlabel="Number of Events", ylabel="Artist")
    dataframes["artists_most_events"] = df3

    # Line Chart
    df4 = run_query(queries["events_by_day"], "events_by_day")
    df4["date"] = df4.apply(lambda x: f"{int(x['year'])}-{int(x['month']):02d}-{int(x['day']):02d}", axis=1)
    generate_chart(df4, "line", "events_by_day.png",
                   x_col="date", y_col="eventcount",
                   title="Event Count by Day",
                   xlabel="Date", ylabel="Number of Events")
    dataframes["events_by_day"] = df4

    # Histogram
    df5 = run_query(queries["rating_distribution"], "rating_distribution")
    generate_chart(df5, "hist", "rating_distribution_hist.png",
                   x_col="rate", y_col="ratingcount",
                   title="Distribution of Event Ratings",
                   xlabel="Rating", ylabel="Frequency")
    dataframes["rating_distribution"] = df5

    # Scatter Plot
    df6 = run_query(queries["rating_vs_user"], "rating_vs_user")
    generate_chart(df6, "scatter", "rating_vs_user.png",
                   x_col="userid", y_col="rate",
                   title="Scatter Plot: Event Ratings vs User ID",
                   xlabel="User ID", ylabel="Rating")
    dataframes["rating_vs_user"] = df6


    df_time = run_query(queries["events_by_date"], "events_by_date")
    plotly_time_slider(df_time)

    export_formatted_excel(dataframes, "events_analytics.xlsx")

if __name__ == "__main__":
    main()
